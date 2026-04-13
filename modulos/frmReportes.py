import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from database.db import getConnection


gnEstadoOperacionCerrada = 2
gnEstadoOperacionPagada = 3


def formatDuration(tnMinutos):
    tnMinutos = int(tnMinutos or 0)
    tnHoras = tnMinutos // 60
    tnMinutosRestantes = tnMinutos % 60

    if tnHoras > 0 and tnMinutosRestantes > 0:
        return f"{tnHoras} h {tnMinutosRestantes} min"

    if tnHoras > 0:
        return f"{tnHoras} h"

    return f"{tnMinutosRestantes} min"


class ReportsView:
    def __init__(self, toParent):
        self.toParent = toParent

        self.lotxtFechaDesde = None
        self.lotxtFechaHasta = None
        self.lotxtPlaca = None
        self.logrdReportes = None
        self.lolblTotalGenerado = None

        self.laFilasActuales = []

    def build(self):
        self.buildFilters()
        self.buildTable()
        self.buildFooter()
        self.loadReports()

    def buildFilters(self):
        lofrmFiltros = tk.Frame(self.toParent, bg="white")
        lofrmFiltros.pack(fill="x", padx=15, pady=15)

        lolblFechaDesde = tk.Label(
            lofrmFiltros,
            text="Desde (YYYY-MM-DD):",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        )
        lolblFechaDesde.grid(row=0, column=0, padx=(5, 5), pady=5, sticky="w")

        self.lotxtFechaDesde = tk.Entry(lofrmFiltros, font=("Arial", 10), width=14)
        self.lotxtFechaDesde.grid(row=0, column=1, padx=(0, 12), pady=5, sticky="w")

        lolblFechaHasta = tk.Label(
            lofrmFiltros,
            text="Hasta (YYYY-MM-DD):",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        )
        lolblFechaHasta.grid(row=0, column=2, padx=(5, 5), pady=5, sticky="w")

        self.lotxtFechaHasta = tk.Entry(lofrmFiltros, font=("Arial", 10), width=14)
        self.lotxtFechaHasta.grid(row=0, column=3, padx=(0, 12), pady=5, sticky="w")

        lolblPlaca = tk.Label(
            lofrmFiltros,
            text="Placa:",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        )
        lolblPlaca.grid(row=0, column=4, padx=(5, 5), pady=5, sticky="w")

        self.lotxtPlaca = tk.Entry(lofrmFiltros, font=("Arial", 10), width=16)
        self.lotxtPlaca.grid(row=0, column=5, padx=(0, 12), pady=5, sticky="w")
        self.lotxtPlaca.bind("<KeyRelease>", lambda toEvent: self.loadReports())

        locmdBuscar = tk.Button(
            lofrmFiltros,
            text="Buscar",
            font=("Arial", 10, "bold"),
            bg="#2563eb",
            fg="white",
            activebackground="#1d4ed8",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.loadReports
        )
        locmdBuscar.grid(row=0, column=6, padx=8, pady=5)

        locmdLimpiar = tk.Button(
            lofrmFiltros,
            text="Limpiar",
            font=("Arial", 10, "bold"),
            bg="#6b7280",
            fg="white",
            activebackground="#4b5563",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.clearFilters
        )
        locmdLimpiar.grid(row=0, column=7, padx=8, pady=5)

        locmdExportar = tk.Button(
            lofrmFiltros,
            text="Exportar Excel",
            font=("Arial", 10, "bold"),
            bg="#16a34a",
            fg="white",
            activebackground="#15803d",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.exportExcel
        )
        locmdExportar.grid(row=0, column=8, padx=8, pady=5)

    def buildTable(self):
        lofrmTabla = tk.Frame(self.toParent, bg="white")
        lofrmTabla.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        taColumnas = (
            "Operacion",
            "CodigoOperacion",
            "Placa",
            "Tarifa",
            "Servicios",
            "Tiempo",
            "FechaIngreso",
            "FechaSalida",
            "MontoParqueo",
            "MontoServicios",
            "MontoTotal"
        )

        self.logrdReportes = ttk.Treeview(
            lofrmTabla,
            columns=taColumnas,
            show="headings",
            height=18
        )

        self.logrdReportes.heading("Operacion", text="ID")
        self.logrdReportes.heading("CodigoOperacion", text="Código")
        self.logrdReportes.heading("Placa", text="Placa")
        self.logrdReportes.heading("Tarifa", text="Tarifa")
        self.logrdReportes.heading("Servicios", text="Servicios")
        self.logrdReportes.heading("Tiempo", text="Tiempo")
        self.logrdReportes.heading("FechaIngreso", text="Fecha ingreso")
        self.logrdReportes.heading("FechaSalida", text="Fecha salida")
        self.logrdReportes.heading("MontoParqueo", text="Parqueo")
        self.logrdReportes.heading("MontoServicios", text="Servicios Bs")
        self.logrdReportes.heading("MontoTotal", text="Monto cobrado")

        self.logrdReportes.column("Operacion", width=60, anchor="center", stretch=False)
        self.logrdReportes.column("CodigoOperacion", width=120, anchor="center", stretch=False)
        self.logrdReportes.column("Placa", width=120, anchor="center", stretch=False)
        self.logrdReportes.column("Tarifa", width=180, anchor="w", stretch=False)
        self.logrdReportes.column("Servicios", width=240, anchor="w", stretch=False)
        self.logrdReportes.column("Tiempo", width=100, anchor="center", stretch=False)
        self.logrdReportes.column("FechaIngreso", width=150, anchor="center", stretch=False)
        self.logrdReportes.column("FechaSalida", width=150, anchor="center", stretch=False)
        self.logrdReportes.column("MontoParqueo", width=100, anchor="center", stretch=False)
        self.logrdReportes.column("MontoServicios", width=110, anchor="center", stretch=False)
        self.logrdReportes.column("MontoTotal", width=120, anchor="center", stretch=False)

        loscrVertical = ttk.Scrollbar(lofrmTabla, orient="vertical", command=self.logrdReportes.yview)
        loscrHorizontal = ttk.Scrollbar(lofrmTabla, orient="horizontal", command=self.logrdReportes.xview)

        self.logrdReportes.configure(
            yscrollcommand=loscrVertical.set,
            xscrollcommand=loscrHorizontal.set
        )

        self.logrdReportes.grid(row=0, column=0, sticky="nsew")
        loscrVertical.grid(row=0, column=1, sticky="ns")
        loscrHorizontal.grid(row=1, column=0, sticky="ew")

        lofrmTabla.grid_rowconfigure(0, weight=1)
        lofrmTabla.grid_columnconfigure(0, weight=1)

    def buildFooter(self):
        lofrmFooter = tk.Frame(self.toParent, bg="white")
        lofrmFooter.pack(fill="x", padx=15, pady=(0, 15))

        self.lolblTotalGenerado = tk.Label(
            lofrmFooter,
            text="Total generado: Bs 0.00",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="#111827"
        )
        self.lolblTotalGenerado.pack(side="right")

    def clearFilters(self):
        self.lotxtFechaDesde.delete(0, tk.END)
        self.lotxtFechaHasta.delete(0, tk.END)
        self.lotxtPlaca.delete(0, tk.END)
        self.loadReports()

    def validateDate(self, tcValor):
        if not tcValor:
            return True

        try:
            datetime.strptime(tcValor, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def getOperationServices(self, loCursor, tnOperacion):
        loCursor.execute("""
            SELECT
                S.Nombre,
                OS.Cantidad
            FROM OPERACIONSERVICIO OS
            INNER JOIN SERVICIO S ON OS.Servicio = S.Servicio
            WHERE OS.Operacion = ?
            ORDER BY S.Nombre ASC
        """, (tnOperacion,))

        laRows = loCursor.fetchall()

        if not laRows:
            return "Parqueo"

        laNombres = []
        for toRow in laRows:
            lcNombre = toRow["Nombre"] or ""
            tnCantidad = int(toRow["Cantidad"] or 0)

            if tnCantidad > 1:
                laNombres.append(f"{lcNombre} x{tnCantidad}")
            else:
                laNombres.append(lcNombre)

        return "Parqueo, " + ", ".join(laNombres)

    def loadReports(self):
        for toItem in self.logrdReportes.get_children():
            self.logrdReportes.delete(toItem)

        lcFechaDesde = self.lotxtFechaDesde.get().strip()
        lcFechaHasta = self.lotxtFechaHasta.get().strip()
        lcPlaca = self.lotxtPlaca.get().strip().upper()

        if not self.validateDate(lcFechaDesde):
            messagebox.showwarning(
                "Fecha inválida",
                "La fecha 'Desde' debe estar en formato YYYY-MM-DD."
            )
            return

        if not self.validateDate(lcFechaHasta):
            messagebox.showwarning(
                "Fecha inválida",
                "La fecha 'Hasta' debe estar en formato YYYY-MM-DD."
            )
            return

        if lcFechaDesde and lcFechaHasta and lcFechaDesde > lcFechaHasta:
            messagebox.showwarning(
                "Rango inválido",
                "La fecha 'Desde' no puede ser mayor que la fecha 'Hasta'."
            )
            return

        loConn = getConnection()
        loCursor = loConn.cursor()

        lcQuery = """
            SELECT
                O.Operacion,
                O.CodigoOperacion,
                V.NumeroPlaca,
                V.LetraPlaca,
                T.Nombre AS NombreTarifa,
                O.FechaIngreso,
                O.FechaSalida,
                O.MontoParqueo,
                O.MontoServicios,
                CAST(
                    (
                        julianday(IFNULL(O.FechaSalida, datetime('now','localtime'))) -
                        julianday(O.FechaIngreso)
                    ) * 24 * 60 AS INTEGER
                ) AS MinutosEstadia
            FROM OPERACION O
            INNER JOIN VEHICULO V ON O.Vehiculo = V.Vehiculo
            INNER JOIN TARIFA T ON O.Tarifa = T.Tarifa
            WHERE O.Estado IN (?, ?)
        """
        laParams = [gnEstadoOperacionCerrada, gnEstadoOperacionPagada]

        if lcFechaDesde:
            lcQuery += " AND date(O.FechaSalida) >= date(?) "
            laParams.append(lcFechaDesde)

        if lcFechaHasta:
            lcQuery += " AND date(O.FechaSalida) <= date(?) "
            laParams.append(lcFechaHasta)

        if lcPlaca:
            lcQuery += """
                AND REPLACE(
                    REPLACE(
                        UPPER(IFNULL(V.NumeroPlaca, '') || IFNULL(V.LetraPlaca, '')),
                        ' ',
                        ''
                    ),
                    '-',
                    ''
                ) LIKE ?
            """
            laParams.append(f"%{lcPlaca.replace(' ', '').replace('-', '')}%")

        lcQuery += " ORDER BY O.FechaSalida DESC, O.Operacion DESC "

        loCursor.execute(lcQuery, laParams)
        laRows = loCursor.fetchall()

        self.laFilasActuales = []
        lnTotalGenerado = 0.0

        for toRow in laRows:
            tnOperacion = toRow["Operacion"]
            lcCodigoOperacion = toRow["CodigoOperacion"] or ""
            lcPlacaFinal = f"{toRow['NumeroPlaca'] or ''} {toRow['LetraPlaca'] or ''}".strip()
            lcTarifa = toRow["NombreTarifa"] or ""
            lcServicios = self.getOperationServices(loCursor, tnOperacion)
            tnMinutos = int(toRow["MinutosEstadia"] or 0)
            lcTiempo = formatDuration(tnMinutos)
            lcFechaIngreso = toRow["FechaIngreso"] or ""
            lcFechaSalida = toRow["FechaSalida"] or ""
            lnMontoParqueo = float(toRow["MontoParqueo"] or 0)
            lnMontoServicios = float(toRow["MontoServicios"] or 0)
            lnMontoTotal = lnMontoParqueo + lnMontoServicios

            self.laFilasActuales.append({
                "Operacion": tnOperacion,
                "CodigoOperacion": lcCodigoOperacion,
                "Placa": lcPlacaFinal,
                "Tarifa": lcTarifa,
                "Servicios": lcServicios,
                "Tiempo": lcTiempo,
                "Minutos": tnMinutos,
                "FechaIngreso": lcFechaIngreso,
                "FechaSalida": lcFechaSalida,
                "MontoParqueo": lnMontoParqueo,
                "MontoServicios": lnMontoServicios,
                "MontoTotal": lnMontoTotal
            })

            self.logrdReportes.insert(
                "",
                "end",
                values=(
                    tnOperacion,
                    lcCodigoOperacion,
                    lcPlacaFinal,
                    lcTarifa,
                    lcServicios,
                    lcTiempo,
                    lcFechaIngreso,
                    lcFechaSalida,
                    f"Bs {lnMontoParqueo:.2f}",
                    f"Bs {lnMontoServicios:.2f}",
                    f"Bs {lnMontoTotal:.2f}"
                )
            )

            lnTotalGenerado += lnMontoTotal

        loConn.close()

        self.lolblTotalGenerado.config(
            text=f"Total generado: Bs {lnTotalGenerado:.2f}"
        )

    def exportExcel(self):
        if not self.laFilasActuales:
            messagebox.showwarning("Sin datos", "No hay datos para exportar.")
            return

        lcRutaArchivo = filedialog.asksaveasfilename(
            title="Guardar reporte",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            initialfile=f"reporte_garaje_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not lcRutaArchivo:
            return

        try:
            loWorkbook = Workbook()
            loSheet = loWorkbook.active
            loSheet.title = "Reportes"

            taHeaders = [
                "ID",
                "Código",
                "Placa",
                "Tarifa",
                "Servicios",
                "Tiempo",
                "Fecha ingreso",
                "Fecha salida",
                "Monto parqueo",
                "Monto servicios",
                "Monto cobrado"
            ]
            loSheet.append(taHeaders)

            for lnColumna in range(1, len(taHeaders) + 1):
                loCelda = loSheet.cell(row=1, column=lnColumna)
                loCelda.font = Font(bold=True)
                loCelda.alignment = Alignment(horizontal="center")

            for txFila in self.laFilasActuales:
                loSheet.append([
                    txFila["Operacion"],
                    txFila["CodigoOperacion"],
                    txFila["Placa"],
                    txFila["Tarifa"],
                    txFila["Servicios"],
                    txFila["Tiempo"],
                    txFila["FechaIngreso"],
                    txFila["FechaSalida"],
                    txFila["MontoParqueo"],
                    txFila["MontoServicios"],
                    txFila["MontoTotal"]
                ])

            lnTotal = sum(txFila["MontoTotal"] for txFila in self.laFilasActuales)
            lnUltimaFila = loSheet.max_row + 2

            loSheet.cell(row=lnUltimaFila, column=10, value="Total generado")
            loSheet.cell(row=lnUltimaFila, column=10).font = Font(bold=True)

            loSheet.cell(row=lnUltimaFila, column=11, value=lnTotal)
            loSheet.cell(row=lnUltimaFila, column=11).font = Font(bold=True)

            gxAnchos = {
                "A": 10,
                "B": 16,
                "C": 14,
                "D": 24,
                "E": 35,
                "F": 14,
                "G": 22,
                "H": 22,
                "I": 16,
                "J": 16,
                "K": 16
            }

            for tcColumna, tnAncho in gxAnchos.items():
                loSheet.column_dimensions[tcColumna].width = tnAncho

            loWorkbook.save(lcRutaArchivo)

            messagebox.showinfo(
                "Exportación exitosa",
                "El reporte se exportó correctamente a Excel."
            )

        except Exception as toError:
            messagebox.showerror(
                "Error",
                f"No se pudo exportar el archivo.\n{str(toError)}"
            )