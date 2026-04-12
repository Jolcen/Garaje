import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from database.db import get_connection


# =========================================================
# CATÁLOGOS
# =========================================================
ESTADO_OPERACION_FINALIZADO = 2
ESTADO_OPERACION_SERVICIO_CANCELADO = 4

TIPO_OPERACION_NORMAL = 1
TIPO_OPERACION_CONTRATO = 2


def nombre_tipo_operacion(tipo_operacion):
    if tipo_operacion == TIPO_OPERACION_CONTRATO:
        return "Contrato"
    return "Normal"


class ReportsView:
    def __init__(self, parent):
        self.parent = parent

        self.entry_from = None
        self.entry_to = None
        self.entry_plate = None
        self.tree = None
        self.total_label = None

        self.current_rows = []

    def build(self):
        self.build_filters()
        self.build_table()
        self.build_footer()
        self.load_reports()

    def build_filters(self):
        filters_frame = tk.Frame(self.parent, bg="white")
        filters_frame.pack(fill="x", padx=15, pady=15)

        tk.Label(
            filters_frame,
            text="Desde (YYYY-MM-DD):",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        ).grid(row=0, column=0, padx=(5, 5), pady=5, sticky="w")

        self.entry_from = tk.Entry(filters_frame, font=("Arial", 10), width=14)
        self.entry_from.grid(row=0, column=1, padx=(0, 12), pady=5, sticky="w")

        tk.Label(
            filters_frame,
            text="Hasta (YYYY-MM-DD):",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        ).grid(row=0, column=2, padx=(5, 5), pady=5, sticky="w")

        self.entry_to = tk.Entry(filters_frame, font=("Arial", 10), width=14)
        self.entry_to.grid(row=0, column=3, padx=(0, 12), pady=5, sticky="w")

        tk.Label(
            filters_frame,
            text="Placa:",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#111827"
        ).grid(row=0, column=4, padx=(5, 5), pady=5, sticky="w")

        self.entry_plate = tk.Entry(filters_frame, font=("Arial", 10), width=14)
        self.entry_plate.grid(row=0, column=5, padx=(0, 12), pady=5, sticky="w")
        self.entry_plate.bind("<KeyRelease>", lambda event: self.load_reports())

        search_button = tk.Button(
            filters_frame,
            text="Buscar",
            font=("Arial", 10, "bold"),
            bg="#2563eb",
            fg="white",
            activebackground="#1d4ed8",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.load_reports
        )
        search_button.grid(row=0, column=6, padx=8, pady=5)

        clear_button = tk.Button(
            filters_frame,
            text="Limpiar",
            font=("Arial", 10, "bold"),
            bg="#6b7280",
            fg="white",
            activebackground="#4b5563",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.clear_filters
        )
        clear_button.grid(row=0, column=7, padx=8, pady=5)

        export_button = tk.Button(
            filters_frame,
            text="Exportar Excel",
            font=("Arial", 10, "bold"),
            bg="#16a34a",
            fg="white",
            activebackground="#15803d",
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.export_excel
        )
        export_button.grid(row=0, column=8, padx=8, pady=5)

    def build_table(self):
        table_frame = tk.Frame(self.parent, bg="white")
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        columns = (
            "Operacion",
            "Placa",
            "Empleado",
            "TipoOperacion",
            "Servicios",
            "Tiempo",
            "FechaIngreso",
            "FechaSalida",
            "MontoParqueo",
            "MontoServicios",
            "MontoTotal"
        )

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=18)

        self.tree.heading("Operacion", text="ID")
        self.tree.heading("Placa", text="Placa")
        self.tree.heading("Empleado", text="Empleado")
        self.tree.heading("TipoOperacion", text="Modalidad")
        self.tree.heading("Servicios", text="Servicios")
        self.tree.heading("Tiempo", text="Tiempo")
        self.tree.heading("FechaIngreso", text="Fecha ingreso")
        self.tree.heading("FechaSalida", text="Fecha salida")
        self.tree.heading("MontoParqueo", text="Parqueo")
        self.tree.heading("MontoServicios", text="Servicios Bs")
        self.tree.heading("MontoTotal", text="Monto cobrado")

        self.tree.column("Operacion", width=60, anchor="center", stretch=False)
        self.tree.column("Placa", width=110, anchor="center", stretch=False)
        self.tree.column("Empleado", width=180, anchor="w", stretch=False)
        self.tree.column("TipoOperacion", width=100, anchor="center", stretch=False)
        self.tree.column("Servicios", width=240, anchor="w", stretch=False)
        self.tree.column("Tiempo", width=100, anchor="center", stretch=False)
        self.tree.column("FechaIngreso", width=150, anchor="center", stretch=False)
        self.tree.column("FechaSalida", width=150, anchor="center", stretch=False)
        self.tree.column("MontoParqueo", width=100, anchor="center", stretch=False)
        self.tree.column("MontoServicios", width=110, anchor="center", stretch=False)
        self.tree.column("MontoTotal", width=120, anchor="center", stretch=False)

        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)

        self.tree.configure(
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

    def build_footer(self):
        footer = tk.Frame(self.parent, bg="white")
        footer.pack(fill="x", padx=15, pady=(0, 15))

        self.total_label = tk.Label(
            footer,
            text="Total generado: Bs 0.00",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="#111827"
        )
        self.total_label.pack(side="right")

    def clear_filters(self):
        self.entry_from.delete(0, tk.END)
        self.entry_to.delete(0, tk.END)
        self.entry_plate.delete(0, tk.END)
        self.load_reports()

    def validate_date(self, value):
        if not value:
            return True
        try:
            datetime.strptime(value, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def format_duration(self, minutes):
        minutes = int(minutes or 0)
        horas = minutes // 60
        mins = minutes % 60

        if horas > 0 and mins > 0:
            return f"{horas} h {mins} min"
        if horas > 0:
            return f"{horas} h"
        return f"{mins} min"

    def load_reports(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        date_from = self.entry_from.get().strip()
        date_to = self.entry_to.get().strip()
        plate = self.entry_plate.get().strip().upper()

        if not self.validate_date(date_from):
            messagebox.showwarning("Fecha inválida", "La fecha 'Desde' debe estar en formato YYYY-MM-DD.")
            return

        if not self.validate_date(date_to):
            messagebox.showwarning("Fecha inválida", "La fecha 'Hasta' debe estar en formato YYYY-MM-DD.")
            return

        if date_from and date_to and date_from > date_to:
            messagebox.showwarning("Rango inválido", "La fecha 'Desde' no puede ser mayor que la fecha 'Hasta'.")
            return

        conn = get_connection()
        cursor = conn.cursor()

        query = """
            SELECT
                O.Operacion,
                V.Placa,
                U.Nombre AS Empleado,
                O.TipoOperacion,
                O.MinutosEstadia,
                O.FechaIngreso,
                O.FechaSalida,
                O.MontoParqueo,
                O.MontoServicios,
                O.MontoTotal
            FROM OPERACION O
            INNER JOIN VEHICULO V ON O.Vehiculo = V.Vehiculo
            LEFT JOIN USUARIO U ON O.UsuarioSalida = U.Usuario
            WHERE O.Estado = ?
        """
        params = [ESTADO_OPERACION_FINALIZADO]

        if date_from:
            query += " AND date(O.FechaSalida) >= date(?)"
            params.append(date_from)

        if date_to:
            query += " AND date(O.FechaSalida) <= date(?)"
            params.append(date_to)

        if plate:
            query += " AND REPLACE(REPLACE(UPPER(V.Placa), ' ', ''), '-', '') LIKE ?"
            params.append(f"%{plate.replace(' ', '').replace('-', '')}%")

        query += " ORDER BY O.FechaSalida DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        self.current_rows = []
        total_generated = 0.0

        for row in rows:
            operation_id = row["Operacion"]
            placa = row["Placa"] or ""
            empleado = row["Empleado"] if row["Empleado"] else "-"
            tipo_operacion = nombre_tipo_operacion(row["TipoOperacion"])
            minutos = int(row["MinutosEstadia"] or 0)
            fecha_ingreso = row["FechaIngreso"] or ""
            fecha_salida = row["FechaSalida"] or ""
            monto_parqueo = float(row["MontoParqueo"] or 0)
            monto_servicios = float(row["MontoServicios"] or 0)
            monto_total = float(row["MontoTotal"] or 0)

            tiempo = self.format_duration(minutos)
            servicios = self.get_operation_services(cursor, operation_id)

            view_row = (
                operation_id,
                placa,
                empleado,
                tipo_operacion,
                servicios,
                tiempo,
                fecha_ingreso,
                fecha_salida,
                f"Bs {monto_parqueo:.2f}",
                f"Bs {monto_servicios:.2f}",
                f"Bs {monto_total:.2f}"
            )

            self.current_rows.append({
                "id": operation_id,
                "placa": placa,
                "empleado": empleado,
                "tipo_operacion": tipo_operacion,
                "servicios": servicios,
                "tiempo": tiempo,
                "minutos": minutos,
                "fecha_ingreso": fecha_ingreso,
                "fecha_salida": fecha_salida,
                "monto_parqueo": monto_parqueo,
                "monto_servicios": monto_servicios,
                "monto_total": monto_total
            })

            self.tree.insert("", "end", values=view_row)
            total_generated += monto_total

        conn.close()

        self.total_label.config(text=f"Total generado: Bs {total_generated:.2f}")

    def get_operation_services(self, cursor, operation_id):
        cursor.execute("""
            SELECT S.Nombre
            FROM OPERACIONSERVICIO OS
            INNER JOIN SERVICIO S ON OS.Servicio = S.Servicio
            WHERE OS.Operacion = ? AND OS.Estado != ?
            ORDER BY S.Nombre ASC
        """, (operation_id, ESTADO_OPERACION_SERVICIO_CANCELADO))
        rows = cursor.fetchall()

        names = [r["Nombre"] for r in rows]
        if not names:
            return "Parqueo"
        return "Parqueo, " + ", ".join(names)

    def export_excel(self):
        if not self.current_rows:
            messagebox.showwarning("Sin datos", "No hay datos para exportar.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Guardar reporte",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            initialfile=f"reporte_garaje_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reportes"

            headers = [
                "ID",
                "Placa",
                "Empleado",
                "Modalidad",
                "Servicios",
                "Tiempo",
                "Fecha ingreso",
                "Fecha salida",
                "Monto parqueo",
                "Monto servicios",
                "Monto cobrado"
            ]
            sheet.append(headers)

            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row in self.current_rows:
                sheet.append([
                    row["id"],
                    row["placa"],
                    row["empleado"],
                    row["tipo_operacion"],
                    row["servicios"],
                    row["tiempo"],
                    row["fecha_ingreso"],
                    row["fecha_salida"],
                    row["monto_parqueo"],
                    row["monto_servicios"],
                    row["monto_total"]
                ])

            total = sum(row["monto_total"] for row in self.current_rows)
            last_row = sheet.max_row + 2
            sheet.cell(row=last_row, column=10, value="Total generado")
            sheet.cell(row=last_row, column=10).font = Font(bold=True)
            sheet.cell(row=last_row, column=11, value=total)
            sheet.cell(row=last_row, column=11).font = Font(bold=True)

            widths = {
                "A": 10,
                "B": 14,
                "C": 22,
                "D": 14,
                "E": 35,
                "F": 14,
                "G": 22,
                "H": 22,
                "I": 16,
                "J": 16,
                "K": 16
            }

            for col_letter, width in widths.items():
                sheet.column_dimensions[col_letter].width = width

            workbook.save(file_path)

            messagebox.showinfo("Exportación exitosa", "El reporte se exportó correctamente a Excel.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el archivo.\n{str(e)}")